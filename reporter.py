import os
import logging
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

# Fill colors
FILL_RED = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFBCC", end_color="FFFBCC", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_RED = Font(color="C0392B", bold=True)
FONT_ORANGE = Font(color="D35400", bold=True)
FONT_YELLOW = Font(color="7D6608", bold=True)


LEVEL_ORDER = {"A级": 0, "B级": 1, "C级": 2, "无": 3}
RISK_ORDER = {"高": 0, "中": 1, "低": 2}


def _sort_key(r: dict):
    risk = RISK_ORDER.get(r.get("complaint_risk", "低"), 2)
    level = LEVEL_ORDER.get(_violation_level(r), 3)
    return (risk, level)


def _should_include(r: dict) -> bool:
    risk = r.get("complaint_risk", "低")
    has_violation = _violation_level(r) != "无"
    return risk in ("高", "中") or has_violation


def _violation_level(r: dict) -> str:
    vs = r.get("violations", [])
    if any(v["level"] == "A级" for v in vs):
        return "A级"
    if any(v["level"] == "B级" for v in vs):
        return "B级"
    if any(v["level"] == "C级" for v in vs):
        return "C级"
    return "无"


_VIOLATION_SHORT = {
    "A1": "引导套奖/弄虚作假", "A2": "泄露客户信息", "A3": "服务态度恶劣",
    "B1": "弄虚作假(首次)", "B2": "泄露机密(首次)", "B3": "严重违纪",
    "B4": "未解答直接转单(投诉)", "B5": "承诺未履行(投诉)", "B6": "超8h未回复",
    "B7": "不合理应对(投诉)",
    "C1": "未解答直接转单", "C2": "承诺未履行", "C3": "超4h未回复",
    "C4": "电话准备不足", "C5": "未回复被挂断", "C6": "不合理应对",
}


def _brief(r: dict) -> str:
    vs = r.get("violations", [])
    if vs:
        parts = [_VIOLATION_SHORT.get(v.get("code", ""), v.get("code", "")) for v in vs]
        return "；".join(parts)
    sig = r.get("complaint_signal")
    if sig:
        return f"投诉信号：{sig[:50]}"
    return r.get("complaint_risk_reason", "")[:50] or ""


def _detail(r: dict) -> str:
    parts = []
    # Attitude
    att = r.get("attitude_score")
    att_issue = r.get("attitude_issues")
    if att is not None:
        parts.append(f"态度：{att}/5" + (f" — {att_issue}" if att_issue else ""))
    # Accuracy
    acc = r.get("accuracy", "")
    acc_issue = r.get("accuracy_issues")
    if acc and acc != "准确":
        parts.append(f"准确性：{acc}" + (f" — {acc_issue}" if acc_issue else ""))
    # Complaint signal
    sig = r.get("complaint_signal")
    if sig:
        parts.append(f"⚠️ 投诉信号：「{sig}」")
    # Violation evidence (brief)
    for v in r.get("violations", []):
        ev = v.get("evidence", "")
        if ev and ev != "系统计算":
            parts.append(f"[{v.get('code','')}] 证据：{ev[:80]}")
    return "\n".join(parts)


def generate_excel(qc_results: list[dict], output_path: str):
    # Filter: only high/medium risk or has violations; sort by risk then level
    sorted_results = sorted(
        [r for r in qc_results if _should_include(r)],
        key=_sort_key
    )

    rows = []
    for r in sorted_results:
        vs = r.get("violations", [])
        rows.append({
            "日期": r.get("date", ""),
            "客服名称": r.get("agent_name", ""),
            "客户ID": r.get("customer_id", ""),
            "违规级别": _violation_level(r),
            "问题简述": _brief(r),
            "详细描述": _detail(r),
            "投诉风险": r.get("complaint_risk", ""),
            "服务态度评分(1-5)": r.get("attitude_score", ""),
            "问题解决情况": r.get("resolution_status", ""),
            "需要跟进": "是" if r.get("action_required") else "否",
            "跟进事项": r.get("action_detail") or "",
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="质检报告", index=False)
        ws = writer.sheets["质检报告"]

        # Style header row
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Color data rows by severity
        for row_idx, row_data in enumerate(rows, start=2):
            level = row_data["违规级别"]
            risk = row_data["投诉风险"]
            fill = None
            if level == "A级":
                fill = FILL_RED
            elif level == "B级" or risk == "高":
                fill = FILL_ORANGE
            elif level == "C级" or risk == "中":
                fill = FILL_YELLOW

            for cell in ws[row_idx]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    cell.fill = fill

            # Bold the violation level cell (column 4)
            level_cell = ws.cell(row=row_idx, column=4)
            if level == "A级":
                level_cell.font = FONT_RED
            elif level == "B级":
                level_cell.font = FONT_ORANGE
            elif level == "C级":
                level_cell.font = FONT_YELLOW

        # Column widths: 日期, 客服, 客户ID, 违规级别, 问题简述, 详细描述, 投诉风险, 态度分, 解决情况, 需跟进, 跟进事项
        col_widths = [12, 14, 14, 10, 45, 70, 10, 14, 14, 10, 50]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    logger.info("Excel 报告已生成: %s", output_path)


def generate_dashboard(qc_results: list[dict], output_path: str, date_str: str):
    total = len(qc_results)
    if total == 0:
        logger.warning("无质检结果，跳过 Dashboard 生成")
        return

    a_count = sum(1 for r in qc_results if _violation_level(r) == "A级")
    b_count = sum(1 for r in qc_results if _violation_level(r) == "B级")
    no_viol = total - a_count - b_count

    high_risk = sum(1 for r in qc_results if r.get("complaint_risk") == "高")
    mid_risk = sum(1 for r in qc_results if r.get("complaint_risk") == "中")
    low_risk = total - high_risk - mid_risk

    avg_att = sum(r.get("attitude_score", 0) or 0 for r in qc_results) / total
    resolved = sum(1 for r in qc_results if r.get("resolution_status") == "已解决")
    res_rate = resolved / total * 100

    action_items = sorted(
        [r for r in qc_results if r.get("action_required") and _should_include(r)],
        key=_sort_key
    )

    # Build action table rows
    action_rows_html = ""
    for r in action_items:
        level = _violation_level(r)
        risk = r.get("complaint_risk", "")
        level_cls = "badge-red" if level == "A级" else ("badge-orange" if level == "B级" else "badge-grey")
        risk_cls = "badge-red" if risk == "高" else ("badge-orange" if risk == "中" else "badge-green")
        brief_text = _brief(r)
        action_rows_html += f"""
        <tr>
          <td>{r.get('agent_name','')}</td>
          <td>{r.get('customer_id','')}</td>
          <td><span class="badge {level_cls}">{level}</span></td>
          <td><span class="badge {risk_cls}">{risk}</span></td>
          <td class="summary-cell">{brief_text}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>WhatsApp 质检日报 {date_str}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
         background: #f0f2f5; color: #333; padding: 20px; }}
  .header {{ background: linear-gradient(135deg, #1a3c5e 0%, #2c5f8a 100%);
             color: #fff; padding: 24px 32px; border-radius: 12px; margin-bottom: 24px; }}
  .header h1 {{ font-size: 22px; font-weight: 700; }}
  .header p {{ margin-top: 6px; opacity: 0.8; font-size: 14px; }}
  .stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 24px; }}
  .stat-card {{ background: #fff; border-radius: 12px; padding: 24px 20px;
                text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,.06); }}
  .stat-num {{ font-size: 42px; font-weight: 800; line-height: 1; }}
  .stat-label {{ margin-top: 8px; font-size: 13px; color: #888; }}
  .red {{ color: #e74c3c; }}
  .orange {{ color: #e67e22; }}
  .blue {{ color: #2980b9; }}
  .green {{ color: #27ae60; }}
  .charts {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px; }}
  .card {{ background: #fff; border-radius: 12px; padding: 20px;
           box-shadow: 0 2px 8px rgba(0,0,0,.06); }}
  .card h3 {{ font-size: 15px; font-weight: 600; margin-bottom: 16px; color: #444; }}
  .chart-wrap {{ position: relative; height: 220px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #1a3c5e; color: #fff; padding: 10px 12px; text-align: left;
        font-weight: 500; white-space: nowrap; }}
  td {{ padding: 10px 12px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }}
  tr:hover td {{ background: #fafafa; }}
  .summary-cell {{ max-width: 300px; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
            font-size: 12px; font-weight: 600; }}
  .badge-red {{ background: #ffe0e0; color: #c0392b; }}
  .badge-orange {{ background: #fff0e0; color: #d35400; }}
  .badge-green {{ background: #e0ffe0; color: #1e8449; }}
  .badge-grey {{ background: #f0f0f0; color: #666; }}
  .progress-bar {{ background: #eee; border-radius: 4px; height: 10px; overflow: hidden; margin-top: 6px; }}
  .progress-fill {{ height: 100%; border-radius: 4px; background: #27ae60; }}
  .metric-row {{ display: flex; justify-content: space-between; align-items: center; padding: 8px 0;
                 border-bottom: 1px solid #f5f5f5; font-size: 14px; }}
  .metric-row:last-child {{ border-bottom: none; }}
  .metric-val {{ font-weight: 700; font-size: 16px; }}
  footer {{ text-align: center; color: #bbb; font-size: 12px; margin-top: 24px; }}
</style>
</head>
<body>

<div class="header">
  <h1>WhatsApp 客服质检日报</h1>
  <p>{date_str} &nbsp;|&nbsp; 共质检 <strong>{total}</strong> 条对话 &nbsp;|&nbsp; 需跟进 <strong>{len(action_items)}</strong> 条</p>
</div>

<div class="stats">
  <div class="stat-card">
    <div class="stat-num red">{a_count}</div>
    <div class="stat-label">A 级违规</div>
  </div>
  <div class="stat-card">
    <div class="stat-num orange">{b_count}</div>
    <div class="stat-label">B 级违规</div>
  </div>
  <div class="stat-card">
    <div class="stat-num orange">{high_risk}</div>
    <div class="stat-label">高投诉风险</div>
  </div>
  <div class="stat-card">
    <div class="stat-num blue">{avg_att:.1f}</div>
    <div class="stat-label">平均态度评分</div>
  </div>
</div>

<div class="charts">
  <div class="card">
    <h3>违规分布</h3>
    <div class="chart-wrap"><canvas id="violChart"></canvas></div>
  </div>
  <div class="card">
    <h3>投诉风险分布</h3>
    <div class="chart-wrap"><canvas id="riskChart"></canvas></div>
  </div>
</div>

<div class="charts">
  <div class="card">
    <h3>关键指标</h3>
    <div class="metric-row">
      <span>问题解决率</span>
      <span class="metric-val green">{res_rate:.0f}%</span>
    </div>
    <div class="progress-bar"><div class="progress-fill" style="width:{res_rate:.0f}%"></div></div>
    <div class="metric-row" style="margin-top:12px">
      <span>违规率（A+B）</span>
      <span class="metric-val {'red' if (a_count+b_count)/total*100 > 20 else 'orange'}">{(a_count+b_count)/total*100:.0f}%</span>
    </div>
    <div class="metric-row">
      <span>高风险占比</span>
      <span class="metric-val {'red' if high_risk/total*100 > 15 else 'orange'}">{high_risk/total*100:.0f}%</span>
    </div>
    <div class="metric-row">
      <span>需跟进 Case</span>
      <span class="metric-val orange">{len(action_items)} / {total}</span>
    </div>
  </div>
  <div class="card">
    <h3>态度评分分布</h3>
    <div class="chart-wrap"><canvas id="attChart"></canvas></div>
  </div>
</div>

<div class="card" style="margin-bottom:24px">
  <h3>需要跟进的 Case（{len(action_items)} 条）</h3>
  <table>
    <tr>
      <th>客服</th><th>客户ID</th><th>违规级别</th><th>投诉风险</th><th>问题描述</th>
    </tr>
    {action_rows_html if action_rows_html else '<tr><td colspan="7" style="text-align:center;color:#888;padding:20px">无需跟进 Case</td></tr>'}
  </table>
</div>

<footer>Generated by moomoo WhatsApp QC System · {date_str}</footer>

<script>
const colors = {{ red:'#e74c3c', orange:'#e67e22', green:'#27ae60', blue:'#3498db', grey:'#95a5a6' }};

new Chart(document.getElementById('violChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['A级违规','B级违规','无违规'],
    datasets: [{{ data: [{a_count},{b_count},{no_viol}],
      backgroundColor: [colors.red, colors.orange, colors.green],
      borderWidth: 2, borderColor: '#fff' }}]
  }},
  options: {{ plugins: {{ legend: {{ position:'bottom' }} }}, cutout:'60%' }}
}});

new Chart(document.getElementById('riskChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['高风险','中风险','低风险'],
    datasets: [{{ data: [{high_risk},{mid_risk},{low_risk}],
      backgroundColor: [colors.red, colors.orange, colors.blue],
      borderWidth: 2, borderColor: '#fff' }}]
  }},
  options: {{ plugins: {{ legend: {{ position:'bottom' }} }}, cutout:'60%' }}
}});

// Attitude score distribution
const attScores = {[r.get('attitude_score',0) or 0 for r in qc_results]};
const attCounts = [1,2,3,4,5].map(s => attScores.filter(x => x===s).length);
new Chart(document.getElementById('attChart'), {{
  type: 'bar',
  data: {{
    labels: ['1分','2分','3分','4分','5分'],
    datasets: [{{ label:'人数', data: attCounts,
      backgroundColor: [colors.red,colors.orange,'#f1c40f',colors.blue,colors.green] }}]
  }},
  options: {{ plugins:{{legend:{{display:false}}}}, scales:{{y:{{beginAtZero:true,ticks:{{stepSize:1}}}}}} }}
}});
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info("Dashboard 已生成: %s", output_path)
